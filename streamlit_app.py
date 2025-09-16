# streamlit_app_ce_batch_persistent.py
import io, os, tempfile, shutil
from dataclasses import dataclass
from typing import List, Dict
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="학회비 일괄자 조회", layout="wide")
st.title("학회비 납부자 일괄 조회")

# -------- Persistent upload cache (disk) --------
UPLOAD_DIR = os.path.join(tempfile.gettempdir(), "dues_cache")
os.makedirs(UPLOAD_DIR, exist_ok=True)

def list_cached_files() -> Dict[str, str]:
    files = {}
    for fn in os.listdir(UPLOAD_DIR):
        path = os.path.join(UPLOAD_DIR, fn)
        if os.path.isfile(path) and fn.lower().endswith(('.xlsx', '.xls')):
            files[fn] = path
    return files

def cache_uploaded_file(uploaded_file) -> str:
    dst = os.path.join(UPLOAD_DIR, uploaded_file.name)
    with open(dst, "wb") as f:
        f.write(uploaded_file.read())
    return dst

def clear_cache():
    if os.path.isdir(UPLOAD_DIR):
        shutil.rmtree(UPLOAD_DIR, ignore_errors=True)
    os.makedirs(UPLOAD_DIR, exist_ok=True)

# -------- Core logic --------
@dataclass
class Hit:
    grade: str
    class_name: str
    row: int
    student_id: str
    name: str
    status_cell: str
    verdict: str  # '납부' or '미납'

def _norm(s):
    if s is None:
        return ""
    return str(s).strip()

def scan_workbook_path(path: str, filename: str, target_id: str,
                       start_row: int = 5, id_col: str = "C", name_col: str = "D", status_col: str = "E") -> List[Hit]:
    wb = load_workbook(path, data_only=True)
    tid = _norm(target_id).replace(" ", "")
    results: List[Hit] = []
    grade = os.path.splitext(filename)[0]  # '1학년.xlsx' -> '1학년'
    for ws in wb.worksheets:
        class_name = ws.title  # A/B/C...
        r = start_row
        while True:
            id_cell = ws[f"{id_col}{r}"].value
            sid = _norm(id_cell)
            if sid == "":
                break  # stop scanning this sheet at first blank in C
            if sid.replace(" ", "") == tid:
                name_val = ws[f"{name_col}{r}"].value
                name_text = _norm(name_val)
                status_val = ws[f"{status_col}{r}"].value
                status_text = _norm(status_val)
                verdict = "미납" if status_text == "미납" else "납부"
                results.append(Hit(grade, class_name, r, sid, name_text, status_text, verdict))
            r += 1
    return results

def batch_check_from_cache(student_ids, start_row=5) -> pd.DataFrame:
    out_rows = []
    cached = list_cached_files()  # {filename: path}
    for sid in student_ids:
        sid = _norm(sid)
        if not sid:
            continue
        found = False
        for fname, fpath in cached.items():
            hits = scan_workbook_path(fpath, fname, sid, start_row=start_row)
            if hits:
                found = True
                for h in hits:
                    out_rows.append({
                        "학번": h.student_id,
                        "이름": h.name,
                        "상태": h.verdict,
                        "원본(E열)": h.status_cell,
                        "학년(파일명)": h.grade,
                        "반(시트명)": h.class_name,
                        "행": h.row
                    })
        if not found:
            out_rows.append({
                "학번": sid,
                "이름": "",
                "상태": "명단에 없음",
                "원본(E열)": "",
                "학년(파일명)": "",
                "반(시트명)": "",
                "행": ""
            })
    return pd.DataFrame(out_rows, columns=["학번","이름","상태","원본(E열)","학년(파일명)","반(시트명)","행"])

# -------- UI --------
with st.sidebar:
    st.header("설정")
    start_row = st.number_input("시작 행", min_value=1, value=5, step=1)
    st.caption("각 시트에서 C열을 이 행부터 내려가며 조회. C열에서 빈 칸을 만나면 그 시트에서 중단.")
    st.header("캐시 관리")
    if st.button("🧹 캐시 비우기(업로드 파일 삭제)"):
        clear_cache()
        st.success("캐시를 비웠습니다. (업로드 파일 삭제)")

st.subheader("1) 학년별 엑셀 파일 업로드 (여러 개 가능, 업로드 후 유지)")
grade_files = st.file_uploader("1학년.xlsx, 2학년.xlsx … 업로드", type=["xlsx"], accept_multiple_files=True)
if grade_files:
    saved = []
    for uf in grade_files:
        path = cache_uploaded_file(uf)
        saved.append(os.path.basename(path))
    if saved:
        st.success(f"업로드 및 캐시 저장 완료: {', '.join(saved)}")

cached = list_cached_files()
if cached:
    st.info(f"현재 캐시된 파일 ({len(cached)}): {', '.join(cached.keys())}")
else:
    st.warning("캐시된 파일이 없습니다. 파일을 업로드하세요.")

st.subheader("2) 학번 목록 업로드 또는 입력")
uploaded_ids = st.file_uploader("학번 CSV/XLSX 파일 (첫 열이 학번)", type=["csv", "xlsx"])
manual_ids = st.text_area("또는 직접 입력 (쉼표/줄바꿈 구분)", placeholder="20230001, 20230002\n20230003")

student_ids = []
if uploaded_ids:
    try:
        if uploaded_ids.name.endswith(".csv"):
            try:
                df = pd.read_csv(uploaded_ids, dtype=str)
            except Exception:
                df = pd.read_csv(uploaded_ids, dtype=str, encoding="cp949")
        else:
            df = pd.read_excel(uploaded_ids, dtype=str)
        student_ids.extend(df.iloc[:, 0].dropna().astype(str).tolist())
    except Exception as e:
        st.error(f"학번 파일을 읽는 중 오류: {e}")

if manual_ids.strip():
    for part in manual_ids.splitlines():
        for x in part.split(","):
            if x.strip():
                student_ids.append(x.strip())

seen = set()
ids_unique = []
for x in student_ids:
    if x not in seen:
        ids_unique.append(x)
        seen.add(x)

if st.button("일괄 조회"):
    if not cached:
        st.warning("먼저 엑셀 파일을 업로드하세요. (업로드 후 캐시에 저장됩니다)")
    elif not ids_unique:
        st.warning("학번을 입력하거나 업로드하세요.")
    else:
        with st.spinner("조회 중…"):
            df_out = batch_check_from_cache(ids_unique, start_row=start_row)
        st.success(f"{len(ids_unique)}건 조회 완료!")
        st.dataframe(df_out, use_container_width=True)
        csv = df_out.to_csv(index=False).encode("utf-8-sig")
        st.download_button("CSV 다운로드", data=csv, file_name="학회비_일괄조회.csv", mime="text/csv")

st.caption("※ 이 앱은 업로드한 엑셀 파일을 서버의 임시 폴더에 보관해 세션이 끊겨도 유지합니다. 단, 클라우드 컨테이너가 재시작되면 삭제될 수 있습니다.")
